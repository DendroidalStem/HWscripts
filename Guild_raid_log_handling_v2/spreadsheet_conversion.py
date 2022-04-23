from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

import os

from pathlib import Path


directory = str(Path(__file__).parent)


def get_color(entry, maxMorale):
    color_type = entry['color_type']
    dic = {
        'UIcorner': 'd9ead4',
        'UI': '94c37f',
        '120': 'ffffff',
        '130': "fffc2e",
        "140": "0bffff",
        '150': '1ffd2c'}

    if color_type in dic:
        return dic[color_type]

    if color_type == 'Morale':
        val = int(entry['value'])
        if val >= maxMorale:
            return "0a803d"
        if val*100 >= maxMorale*95:
            return "25b864"
        if val*100 >= maxMorale*90:
            return "4ef596"
        if val*100 >= maxMorale*80:
            return "ffd66c"
        if val*100 >= maxMorale*65:
            return "ff9519"
        else:
            return "ff0006"

# ffd66c dark green

# 3a7523 mid green

# ffd66c yellowish

# 3a7523 light green

# ff9519 orangeish

# ff0006 red

    return "ffffff"


def create_spreadsheet(spreadsheet_data, maxMorale, output_file):
    wb = Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i, row in enumerate(spreadsheet_data):
        for j, entry in enumerate(row):
            val = entry['value']
            color = get_color(entry, maxMorale)
            if val:
                ws.cell(row=i+1, column=j+1, value=val)
            ws.cell(row=i+1, column=j+1).fill = PatternFill(patternType='solid',
                                                            fgColor=color)
            ws.cell(row=i+1, column=j+1).alignment = Alignment(horizontal="center")
            ws.cell(row=i+1, column=j+1).border = thin_border
            if type(val) == int:
                ws.cell(row=i+1, column=j+1).number_format = '#,###'

    # Save the file
    print(os.path.join(directory, output_file))
    wb.save(os.path.join(directory, output_file))


#############
