from openpyxl import Workbook
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.styles.alignment import Alignment

import os

from pathlib import Path


directory = str(Path(__file__).parent)


def get_color(cell_type):
    dic = {
        'UIcorner': 'd9ead4',
        'UI': '94c37f',
        '120': 'ffffff',
        '130': "fffc2e",
        "140": "0bffff",
        '150': '1ffd2c'}
    return dic[cell_type] if cell_type in dic else "ffffff"


def create_spreadsheet(spreadsheet_data, output_file):
    wb = Workbook()
    ws = wb.active

    thin_border = Border(left=Side(style='thin'),
                         right=Side(style='thin'),
                         top=Side(style='thin'),
                         bottom=Side(style='thin'))

    for i, row in enumerate(spreadsheet_data):
        for j, entry in enumerate(row):
            val = entry['value']
            color = get_color(entry['color_type'])
            if val:
                ws.cell(row=i+1, column=j+1, value=val)
            ws.cell(row=i+1, column=j+1).fill = PatternFill(patternType='solid',
                                                            fgColor=color)
            ws.cell(row=i+1, column=j+1).alignment = Alignment(horizontal="center")
            ws.cell(row=i+1, column=j+1).border = thin_border
            if type(val) == int:
                ws.cell(row=i+1, column=j+1).number_format = '#,###'

    # Save the file
    print(os.path.join(directory, output_file))
    wb.save(os.path.join(directory, output_file))
